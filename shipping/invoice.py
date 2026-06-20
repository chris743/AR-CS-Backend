"""Combine the shipping report with phyto matches, and render the invoice xlsx.

Ports combine_reports.py + output_xlsx.py. pandas + openpyxl are base deps.
"""

import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

from . import config

_WE_RE = re.compile(config.WE_FOLDER_REGEX)
_BLUE = Font(color="0000FF")

_TABLE_HEADERS = ["SHIP DATE", "ORDER #", "MATERIAL", "QTY", "COST EACH", "TOTAL"]
_HEADER_ROW = 8
_DATA_START_ROW = 9
_TOTAL_COL_IDX = len(_TABLE_HEADERS)


def combine_reports(trimmed_csv_path: str, phyto_results: list[dict],
                    output_path: str = config.COMBINED_LOCAL) -> str:
    """Left-join shipping rows (sono) with phyto match data (order_number)."""
    ship = pd.read_csv(trimmed_csv_path)

    sono_col = next((c for c in ship.columns if c.lower() == "sono"), None)
    if sono_col is None:
        raise KeyError(f"'sono' column not found in {trimmed_csv_path}. Got: {list(ship.columns)}")

    phyto_rows = []
    for r in phyto_results:
        m = r.get("match", {}) or {}
        phyto_rows.append({
            "order_number": r.get("order_number"),
            "certificate_number": m.get("current_certificate_number"),
            "debit_date": m.get("date"),
            "debit_amount": m.get("debit_amount"),
            "phyto_matched": m.get("matched"),
        })
    phyto = pd.DataFrame(phyto_rows)

    ship[sono_col] = pd.to_numeric(ship[sono_col], errors="coerce")
    if not phyto.empty:
        phyto["order_number"] = pd.to_numeric(phyto["order_number"], errors="coerce")
        combined = ship.merge(phyto, how="left", left_on=sono_col, right_on="order_number")
        combined = combined.drop(columns=["order_number"])
    else:
        combined = ship

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    combined.to_csv(output_path, index=False)
    return output_path


def _week_ending_str(week_folder: str) -> str:
    m = _WE_RE.match(week_folder or "")
    if not m:
        return ""
    mm, dd, yyyy = m.groups()
    return f"{mm}/{dd}/{yyyy}"


def _format_ship_date(value) -> str:
    d = pd.to_datetime(value, errors="coerce")
    if pd.isna(d):
        return "" if pd.isna(value) else str(value)
    return d.strftime("%m/%d/%Y")


def build_xlsx(combined_csv_path: str, week_folder: str,
               phyto_results: list[dict] | None = None,
               output_path: str | None = None) -> str:
    """Write a formatted invoice xlsx from the combined report plus phyto matches."""
    if output_path is None:
        Path(config.XLSX_DIR).mkdir(parents=True, exist_ok=True)
        safe = (week_folder or "creekside").replace(" ", "_").replace(".", "-")
        output_path = str(Path(config.XLSX_DIR) / f"{safe}.xlsx")

    df = pd.read_csv(combined_csv_path)
    phyto_results = phyto_results or []

    wb = Workbook()
    ws = wb.active
    ws.title = "Charges"

    ws["A1"] = "COBBLESTONE FRUIT"
    ws["A2"] = "CREEKSIDE MATERIALS CHARGES"
    ws["A3"], ws["B3"] = "CUSTID:", config.INVOICE_CUSTID
    ws["A4"], ws["B4"] = "WEEK ENDING", _week_ending_str(week_folder)
    ws["A5"], ws["B5"] = "CHARGE CODE:", config.INVOICE_CHARGE_CODE
    ws["A6"] = "INVOICE#:"

    for col_idx, header in enumerate(_TABLE_HEADERS, start=1):
        ws.cell(row=_HEADER_ROW, column=col_idx, value=header)

    for offset, row in enumerate(df.itertuples(index=False)):
        r = _DATA_START_ROW + offset
        ws.cell(row=r, column=1, value=_format_ship_date(row.shipdatetime))
        ws.cell(row=r, column=2, value=row.sono)
        ws.cell(row=r, column=3, value=row.chargedescr)
        ws.cell(row=r, column=4, value=row.qnt)
        ws.cell(row=r, column=5, value=row.rate)
        ws.cell(row=r, column=6, value=row.amt)

    phyto_start = _DATA_START_ROW + len(df)
    for offset, result in enumerate(phyto_results):
        r = phyto_start + offset
        match = result.get("match") or {}
        amount = match.get("debit_amount")
        ws.cell(row=r, column=1, value=_format_ship_date(match.get("date")))
        ws.cell(row=r, column=2, value=result.get("order_number"))
        ws.cell(row=r, column=3, value="phytosanitary inspection")
        ws.cell(row=r, column=4, value=1)
        ws.cell(row=r, column=5, value=amount)
        ws.cell(row=r, column=6, value=amount)

    last_row = _DATA_START_ROW + len(df) + len(phyto_results) - 1
    for col_idx in range(1, _TOTAL_COL_IDX):
        for row_idx in range(_HEADER_ROW, last_row + 1):
            ws.cell(row=row_idx, column=col_idx).font = _BLUE

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
